import os
import csv
import io
from datetime import datetime, timedelta

# Optional import for Excel support
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    jsonify,
    Response,
    current_app,
)
from werkzeug.utils import secure_filename
from bson.objectid import ObjectId

from . import manager_bp
from extensions import mongo


# =========================================================
# FILE UPLOAD HELPER (for profile picture)
# =========================================================
def allowed_file(filename):
    """
    Check if the uploaded file has an allowed extension.
    Uses ALLOWED_EXTENSIONS from app config.
    """
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower()
        in current_app.config.get("ALLOWED_EXTENSIONS", set())
    )


# =========================================================
# AUTH HELPERS
# =========================================================
def login_required(f):
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login first.", "warning")
            return redirect("/login")
        return f(*args, **kwargs)

    return decorated


def manager_required(f):
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session or session.get("role") != "manager":
            flash("Manager access required.", "danger")
            return redirect("/login")
        return f(*args, **kwargs)

    return decorated


# =========================================================
# SHIFT HELPERS
# =========================================================
def check_shift_conflict(user_id, date_str, exclude_shift_id=None):
    """
    Check if a user already has a shift on the given date.
    Returns (has_conflict, existing_shift) tuple.
    """
    query = {
        "user_id": ObjectId(user_id) if isinstance(user_id, str) else user_id,
        "date": date_str,
    }
    existing_shift = mongo.db.shifts.find_one(query)

    if existing_shift:
        # If we're updating an existing shift, exclude it from conflict check
        if exclude_shift_id and str(existing_shift["_id"]) == str(exclude_shift_id):
            return False, None
        return True, existing_shift

    return False, None


# Color map for calendar events by shift code
SHIFT_COLORS = {
    "A": "#0d6efd",  # blue
    "B": "#198754",  # green
    "C": "#ffc107",  # yellow
    "G": "#6f42c1",  # purple
    "W": "#6c757d",  # gray (weekoff)
    "L": "#dc3545",  # red (leave)
}

# OFFICIAL SHIFT TIMINGS (24-hour format)
SHIFT_TIMINGS = {
    "A": ("06:00", "14:30"),  # 6 AM → 2:30 PM
    "B": ("14:00", "22:30"),  # 2 PM → 10:30 PM
    "C": ("22:00", "06:00"),  # 10 PM → 6 AM NEXT DAY
    "G": ("09:00", "17:30"),  # 9 AM → 5:30 PM
    "W": ("00:00", "23:59"),  # Weekoff - all day
    "L": ("00:00", "23:59"),  # Leave - all day
}


# =========================================================
# MANAGER DASHBOARD
# =========================================================
@manager_bp.route("/dashboard", methods=["GET"], endpoint="dashboard")
@manager_required
def dashboard():
    users_count = mongo.db.users.count_documents({})
    shifts_count = mongo.db.shifts.count_documents({})
    project_count = mongo.db.projects.count_documents({})
    pending_change = mongo.db.shift_change_requests.count_documents(
        {"status": "pending"}
    )
    pending_swap = mongo.db.shift_swap_requests.count_documents({"status": "pending"})

    projects = list(mongo.db.projects.find().sort("created_at", -1))
    for p in projects:
        p["shift_count"] = mongo.db.shifts.count_documents({"project_id": p["_id"]})
        p["task_count"] = mongo.db.project_tasks.count_documents({"project_id": p["_id"]})

    return render_template(
        "manager/dashboard.html",
        users_count=users_count,
        shifts_count=shifts_count,
        project_count=project_count,
        pending_change=pending_change,
        pending_swap=pending_swap,
        projects=projects,
    )


# =========================================================
# AUTO ROSTER (USES OFFICIAL SHIFT TIMINGS)
# =========================================================
@manager_bp.route("/auto-roster", methods=["POST"], endpoint="auto_roster")
@manager_required
def auto_roster():
    start_date_str = request.form.get("start_date")
    end_date_str = request.form.get("end_date")
    project_id = request.form.get("project_id") or None

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        flash("Invalid date range", "danger")
        return redirect(url_for("manager.dashboard"))

    if end_date < start_date:
        flash("End date must be after start date", "danger")
        return redirect(url_for("manager.dashboard"))

    members = list(mongo.db.users.find({"role": "member"}))
    if not members:
        flash("No team members found to assign shifts.", "warning")
        return redirect(url_for("manager.dashboard"))

    member_ids = [m["_id"] for m in members]
    idx = 0

    day = start_date
    while day <= end_date:
        date_str = day.isoformat()

        # Loop over each shift code (A, B, C, G)
        for shift_code, (start_t, end_t) in SHIFT_TIMINGS.items():
            user_id = member_ids[idx % len(member_ids)]
            idx += 1

            existing_shift = mongo.db.shifts.find_one(
                {"date": date_str, "user_id": user_id}
            )

            doc = {
                "date": date_str,
                "user_id": user_id,
                "shift_code": shift_code,
                "start_time": start_t,
                "end_time": end_t,
                "task": f"{shift_code} shift",
                "updated_at": datetime.utcnow(),
            }
            if project_id:
                doc["project_id"] = ObjectId(project_id)
            else:
                doc["project_id"] = None

            if existing_shift:
                mongo.db.shifts.update_one(
                    {"_id": existing_shift["_id"]},
                    {"$set": doc},
                )
            else:
                doc["created_at"] = datetime.utcnow()
                mongo.db.shifts.insert_one(doc)

        day += timedelta(days=1)

    flash("Auto roster generated successfully.", "success")
    return redirect(url_for("manager.dashboard"))


# =========================================================
# API: SHIFTS FOR MANAGER CALENDAR
# =========================================================
@manager_bp.route("/api/shifts", endpoint="api_shifts")
@manager_required
def api_shifts():
    """
    Returns shifts as FullCalendar events (for manager calendar).
    Optional filters: ?user_id=...&project_id=...
    Production-ready with error handling and caching considerations.
    """
    try:
        query = {}
        user_id = request.args.get("user_id")
        project_id = request.args.get("project_id")

        if user_id:
            try:
                query["user_id"] = ObjectId(user_id)
            except:
                return jsonify({"error": "Invalid user_id"}), 400
        if project_id:
            try:
                query["project_id"] = ObjectId(project_id)
            except:
                return jsonify({"error": "Invalid project_id"}), 400

        shifts_cursor = mongo.db.shifts.find(query).sort("date", 1)

        users_list = list(mongo.db.users.find())
        users_map = {str(u["_id"]): {"name": u.get("name", "Unknown"), "profile_pic": u.get("profile_picture", "default.png")} for u in users_list}
        projects_map = {str(p["_id"]): p["name"] for p in mongo.db.projects.find()}

        events = []
        for s in shifts_cursor:
            try:
                date_str = s.get("date")  # "YYYY-MM-DD"
                if not date_str:
                    continue
                    
                start_time = s.get("start_time") or "09:00"
                end_time = s.get("end_time") or "17:00"
                shift_code = s.get("shift_code", "")

                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                except:
                    continue

                start = f"{date_str}T{start_time}:00"

                # Night shift: C → end next day
                # Weekoff/Leave: all day events
                if shift_code == "C":
                    end_date = date_obj + timedelta(days=1)
                    end = f"{end_date.isoformat()}T{end_time}:00"
                elif shift_code in ["W", "L"]:
                    # All day event for weekoff/leave
                    end = f"{date_str}T{end_time}:00"
                else:
                    end = f"{date_str}T{end_time}:00"

                uid = str(s.get("user_id", ""))
                pid = str(s.get("project_id")) if s.get("project_id") else None

                user_info = users_map.get(uid, {"name": "Unknown", "profile_pic": "default.png"})
                user_name = user_info["name"]
                profile_pic = user_info["profile_pic"]
                project_name = projects_map.get(pid, "General") if pid else "General"
                task = s.get("task", "")

                color = SHIFT_COLORS.get(shift_code, "#0dcaf0")
                shift_label = "Weekoff" if shift_code == "W" else ("Leave" if shift_code == "L" else shift_code)
                tooltip = f"{user_name} • {project_name} • {shift_label} • {task}"

                shift_label = "Weekoff" if shift_code == "W" else ("Leave" if shift_code == "L" else shift_code)
                events.append(
                    {
                        "id": str(s["_id"]),
                        "title": f"{project_name}: {user_name} – {shift_label}",
                        "start": start,
                        "end": end,
                        "backgroundColor": color,
                        "borderColor": color,
                        "allDay": shift_code in ["W", "L"],  # All day for weekoff/leave
                        "extendedProps": {
                            "shift_id": str(s["_id"]),
                            "user": user_name,
                            "user_id": uid,
                            "profile_pic": profile_pic,
                            "project": project_name,
                            "task": task,
                            "shift_code": shift_code,
                            "tooltip": tooltip,
                        },
                    }
                )
            except Exception as e:
                # Log error but continue processing other shifts
                current_app.logger.error(f"Error processing shift {s.get('_id')}: {str(e)}")
                continue

        return jsonify(events)
    except Exception as e:
        current_app.logger.error(f"Error in api_shifts: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500


# =========================================================
# API: CALENDAR DRAG/RESIZE UPDATE
# =========================================================
@manager_bp.route("/api/update-shift", methods=["POST"], endpoint="api_update_shift")
@manager_required
def api_update_shift():
    """
    Update shift date/time when dragged or resized on the calendar.
    (Keeps whatever times user drags to; does not snap back to template.)
    """
    data = request.get_json(force=True)
    shift_id = data.get("id")
    start_iso = data.get("start")
    end_iso = data.get("end")

    if not shift_id or not start_iso or not end_iso:
        return jsonify({"success": False, "error": "Missing data"}), 400

    try:
        start_dt = datetime.fromisoformat(start_iso.replace("Z", "+00:00"))
        end_dt = datetime.fromisoformat(end_iso.replace("Z", "+00:00"))
    except ValueError:
        return jsonify({"success": False, "error": "Invalid datetime"}), 400

    date_str = start_dt.date().isoformat()
    start_time = start_dt.strftime("%H:%M")
    end_time = end_dt.strftime("%H:%M")

    current_shift = mongo.db.shifts.find_one({"_id": ObjectId(shift_id)})
    if not current_shift:
        return jsonify({"success": False, "error": "Shift not found"}), 404

    if current_shift.get("date") != date_str:
        has_conflict, _ = check_shift_conflict(
            current_shift["user_id"], date_str, exclude_shift_id=shift_id
        )
        if has_conflict:
            return jsonify(
                {
                    "success": False,
                    "error": f"User already has a shift on {date_str}. Cannot move shift to this date.",
                }
            ), 400

    mongo.db.shifts.update_one(
        {"_id": ObjectId(shift_id)},
        {
            "$set": {
                "date": date_str,
                "start_time": start_time,
                "end_time": end_time,
                "updated_at": datetime.utcnow(),
            }
        },
    )

    return jsonify({"success": True})


# =========================================================
# TEAM VIEW (MANAGER – PROJECT FILTER)
# =========================================================
@manager_bp.route("/project-members", endpoint="project_members")
@manager_required
def project_members():
    selected_project = request.args.get("project_id")
    selected_date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")

    # Get all projects and convert IDs to string
    projects_raw = list(mongo.db.projects.find().sort("name", 1))
    projects = []
    for p in projects_raw:
        p["_id"] = str(p["_id"])
        projects.append(p)

    members_data = []

    if selected_project:
        try:
            pid = ObjectId(selected_project)
            # Get all members allocated to this project
            members = list(
                mongo.db.users.find(
                    {"role": "member", "project_ids": pid}
                )
            )
            
            # Get shifts for this project and date range (current week)
            from datetime import timedelta
            start_date = datetime.strptime(selected_date, "%Y-%m-%d")
            end_date = start_date + timedelta(days=6)
            start_str = start_date.strftime("%Y-%m-%d")
            end_str = end_date.strftime("%Y-%m-%d")
            
            # Get all shifts for this project in the date range
            shifts = list(mongo.db.shifts.find({
                "project_id": pid,
                "date": {"$gte": start_str, "$lte": end_str}
            }).sort("date", 1))
            
            # Get project tasks
            tasks = list(mongo.db.project_tasks.find({
                "project_id": pid
            }))
            
            # Organize data by member
            members_map = {str(m["_id"]): m for m in members}
            for member in members:
                member_id = str(member["_id"])
                member["_id"] = member_id
                
                # Get shifts for this member in this project
                member_shifts = []
                for s in shifts:
                    if str(s.get("user_id")) == member_id:
                        # Convert ObjectIds to strings for template
                        s["_id"] = str(s["_id"])
                        if s.get("user_id"):
                            s["user_id"] = str(s["user_id"])
                        if s.get("project_id"):
                            s["project_id"] = str(s["project_id"])
                        member_shifts.append(s)
                
                # Get tasks assigned to this member
                member_tasks = []
                for t in tasks:
                    if str(t.get("assigned_to")) == member_id:
                        # Convert ObjectIds to strings for template
                        t["_id"] = str(t["_id"])
                        if t.get("project_id"):
                            t["project_id"] = str(t["project_id"])
                        if t.get("assigned_to"):
                            t["assigned_to"] = str(t["assigned_to"])
                        member_tasks.append(t)
                
                members_data.append({
                    "member": member,
                    "shifts": member_shifts,
                    "tasks": member_tasks,
                    "shift_count": len(member_shifts),
                    "task_count": len(member_tasks)
                })
        except Exception as e:
            selected_project = None

    return render_template(
        "manager/project_members.html",
        members_data=members_data,
        projects=projects,
        selected_project=selected_project,
        selected_date=selected_date,
    )


# =========================================================
# SHIFT-WISE VIEW (GROUP BY SHIFT CODE)
# =========================================================
@manager_bp.route("/shift-wise-view", endpoint="shift_wise_view")
@manager_required
def shift_wise_view():
    selected_date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    selected_project = request.args.get("project_id")
    sort_alphabetical = request.args.get("sort") == "alphabetical"
    
    # Build query
    query = {"date": selected_date}
    if selected_project:
        query["project_id"] = ObjectId(selected_project)
    
    # Get all shifts for the selected date
    shifts = list(mongo.db.shifts.find(query))
    
    # Get all users and projects for mapping
    all_users = list(mongo.db.users.find())
    users_map = {str(u["_id"]): u for u in all_users}
    projects_map = {str(p["_id"]): p["name"] for p in mongo.db.projects.find()}
    
    # Group shifts by shift code
    shift_groups = {
        "A": [],
        "B": [],
        "C": [],
        "G": []
    }
    
    for shift in shifts:
        shift_code = shift.get("shift_code", "")
        if shift_code in shift_groups:
            user_id = str(shift.get("user_id"))
            user = users_map.get(user_id)
            if user:
                shift_info = {
                    "shift_id": str(shift["_id"]),
                    "user_id": user_id,
                    "user_name": user.get("name", "Unknown"),
                    "user_email": user.get("email", ""),
                    "project_id": str(shift.get("project_id")) if shift.get("project_id") else None,
                    "project_name": projects_map.get(str(shift.get("project_id")), "General") if shift.get("project_id") else "General",
                    "task": shift.get("task", ""),
                    "start_time": shift.get("start_time", ""),
                    "end_time": shift.get("end_time", "")
                }
                shift_groups[shift_code].append(shift_info)
    
    # Sort alphabetically if requested
    if sort_alphabetical:
        for shift_code in shift_groups:
            shift_groups[shift_code].sort(key=lambda x: x["user_name"].lower())
    
    # Get counts
    shift_counts = {code: len(members) for code, members in shift_groups.items()}
    total_members = sum(shift_counts.values())
    
    # Get all projects for filter
    projects = list(mongo.db.projects.find().sort("name", 1))
    for p in projects:
        p["_id"] = str(p["_id"])
    
    # Get today's date for date input
    from datetime import date
    today_date = date.today().isoformat()
    
    return render_template(
        "manager/shift_wise_view.html",
        shift_groups=shift_groups,
        shift_counts=shift_counts,
        total_members=total_members,
        selected_date=selected_date,
        selected_project=selected_project,
        projects=projects,
        sort_alphabetical=sort_alphabetical,
        today_date=today_date,
    )


# =========================================================
# EXPORT CSV
# =========================================================
@manager_bp.route("/export/csv", endpoint="export_csv")
@manager_required
def export_csv():
    project_id = request.args.get("project_id")
    query = {}
    if project_id:
        query["project_id"] = ObjectId(project_id)

    shifts = mongo.db.shifts.find(query).sort("date", 1)
    users_map = {str(u["_id"]): u["name"] for u in mongo.db.users.find()}
    projects_map = {str(p["_id"]): p["name"] for p in mongo.db.projects.find()}

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "User", "Project", "Shift", "Start", "End", "Task"])

    for s in shifts:
        uid = str(s["user_id"])
        pid = str(s.get("project_id")) if s.get("project_id") else None
        writer.writerow(
            [
                s["date"],
                users_map.get(uid, "Unknown"),
                projects_map.get(pid, "General") if pid else "General",
                s.get("shift_code", ""),
                s.get("start_time", ""),
                s.get("end_time", ""),
                s.get("task", ""),
            ]
        )

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=shift_schedule.csv"},
    )


# =========================================================
# EXPORT PRINT VIEW
# =========================================================
@manager_bp.route("/export/print", endpoint="export_print")
@manager_required
def export_print():
    project_id = request.args.get("project_id")
    query = {}
    if project_id:
        query["project_id"] = ObjectId(project_id)

    shifts = list(mongo.db.shifts.find(query).sort("date", 1))
    users_map = {str(u["_id"]): u["name"] for u in mongo.db.users.find()}
    projects_map = {str(p["_id"]): p["name"] for p in mongo.db.projects.find()}

    return render_template(
        "manager/export_print.html",
        shifts=shifts,
        users_map=users_map,
        projects_map=projects_map,
    )


# =========================================================
# MANAGE SHIFTS (MANUAL ADD + TASKS, AUTO TIME)
# =========================================================
@manager_bp.route("/manage-shifts", methods=["GET", "POST"], endpoint="manage_shifts")
@manager_required
def manage_shifts():
    projects = list(mongo.db.projects.find())
    selected_date = request.args.get("date")
    selected_project = request.args.get("project_id")
    show_all_members = request.args.get("show_all", "false") == "true"

    if selected_project and not show_all_members:
        users = list(
            mongo.db.users.find(
                {"role": "member", "project_ids": ObjectId(selected_project)}
            )
        )
    else:
        users = list(mongo.db.users.find({"role": "member"}))

    if request.method == "POST":
        action = request.form.get("action")

        # ---------------- BULK SHIFT ASSIGNMENT ----------------
        if action == "bulk_add_shifts":
            project_id = request.form.get("bulk_project_id")
            shift_code = request.form.get("bulk_shift_code")
            date_str = request.form.get("bulk_date")
            task = request.form.get("bulk_task", "")
            selected_members = request.form.getlist("selected_members")
            
            if not project_id or not shift_code or not date_str:
                flash("Please fill in all required fields (Project, Shift, Date).", "danger")
                redirect_url = url_for("manager.manage_shifts")
                if selected_project:
                    redirect_url += f"?project_id={selected_project}"
                return redirect(redirect_url)
            
            if not selected_members:
                flash("Please select at least one member.", "danger")
                redirect_url = url_for("manager.manage_shifts")
                if selected_project:
                    redirect_url += f"?project_id={selected_project}"
                return redirect(redirect_url)
            
            start_time, end_time = SHIFT_TIMINGS.get(shift_code, ("09:00", "17:00"))
            success_count = 0
            conflict_count = 0
            skipped_count = 0
            
            for user_id in selected_members:
                # Check for existing shift
                existing_shift = mongo.db.shifts.find_one({
                    "date": date_str,
                    "user_id": ObjectId(user_id)
                })
                
                if existing_shift:
                    conflict_count += 1
                    continue
                
                # Create shift
                doc = {
                    "date": date_str,
                    "user_id": ObjectId(user_id),
                    "shift_code": shift_code,
                    "start_time": start_time,
                    "end_time": end_time,
                    "task": task,
                    "project_id": ObjectId(project_id),
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow()
                }
                
                mongo.db.shifts.insert_one(doc)
                
                # Send notification
                mongo.db.notifications.insert_one({
                    "user_id": ObjectId(user_id),
                    "message": f"You have been assigned a {shift_code} shift on {date_str}.",
                    "created_at": datetime.utcnow(),
                    "read": False,
                })
                
                success_count += 1
            
            # Flash message with results
            if success_count > 0:
                flash(f"Successfully assigned shifts to {success_count} member(s)!", "success")
            if conflict_count > 0:
                flash(f"{conflict_count} member(s) already had shifts on this date and were skipped.", "warning")
            
            redirect_url = url_for("manager.manage_shifts", date=date_str)
            if selected_project:
                redirect_url += f"&project_id={selected_project}"
            return redirect(redirect_url)

        # ---------------- TASK CREATION ----------------
        if action == "add_task":
            task_name = request.form.get("task_name")
            assigned_to = request.form.get("assigned_to")
            due_date = request.form.get("due_date")
            project_id = request.form.get("project_id")

            if not project_id:
                flash("Please select a project.", "danger")
                return redirect(url_for("manager.manage_shifts"))

            mongo.db.project_tasks.insert_one(
                {
                    "project_id": ObjectId(project_id),
                    "task_name": task_name,
                    "assigned_to": ObjectId(assigned_to),
                    "due_date": due_date,
                    "created_at": datetime.utcnow(),
                }
            )

            mongo.db.notifications.insert_one(
                {
                    "user_id": ObjectId(assigned_to),
                    "message": f"New task '{task_name}' assigned to you for project.",
                    "project_id": ObjectId(project_id),
                    "created_at": datetime.utcnow(),
                    "read": False,
                }
            )

            flash("Task added successfully!", "success")
            redirect_url = url_for("manager.manage_shifts")
            if selected_project:
                redirect_url += f"?project_id={selected_project}"
            return redirect(redirect_url)

        # ---------------- SHIFT CREATION / UPDATE (AUTO TIME) ----------------
        if action == "add_shift" or action is None:
            date_str = request.form.get("date")
            user_id = request.form.get("user_id")
            shift_code = request.form.get("shift_code")
            task = request.form.get("task")
            project_id = request.form.get("project_id") or None

            if not date_str or not user_id or not shift_code:
                flash("Please fill in all required fields for shift.", "danger")
                redirect_url = url_for("manager.manage_shifts")
                if selected_project:
                    redirect_url += f"?project_id={selected_project}"
                return redirect(redirect_url)

            start_time, end_time = SHIFT_TIMINGS.get(
                shift_code, ("09:00", "17:00")
            )

            existing_shift = mongo.db.shifts.find_one(
                {"date": date_str, "user_id": ObjectId(user_id)}
            )

            doc = {
                "date": date_str,
                "user_id": ObjectId(user_id),
                "shift_code": shift_code,
                "start_time": start_time,
                "end_time": end_time,
                "task": task,
                "updated_at": datetime.utcnow(),
            }

            if project_id:
                doc["project_id"] = ObjectId(project_id)
            else:
                doc["project_id"] = None

            if existing_shift:
                mongo.db.shifts.update_one(
                    {"_id": existing_shift["_id"]}, {"$set": doc}
                )
                mongo.db.notifications.insert_one(
                    {
                        "user_id": ObjectId(user_id),
                        "message": f"Your shift on {date_str} has been updated.",
                        "created_at": datetime.utcnow(),
                        "read": False,
                    }
                )
                flash("Shift updated successfully.", "success")
            else:
                doc["created_at"] = datetime.utcnow()
                mongo.db.shifts.insert_one(doc)
                mongo.db.notifications.insert_one(
                    {
                        "user_id": ObjectId(user_id),
                        "message": f"You have been assigned a shift on {date_str}.",
                        "created_at": datetime.utcnow(),
                        "read": False,
                    }
                )
                flash("Shift created successfully.", "success")

            redirect_url = url_for("manager.manage_shifts", date=date_str)
            if selected_project:
                redirect_url += f"&project_id={selected_project}"
            return redirect(redirect_url)

    query = {}
    if selected_date:
        query["date"] = selected_date
    if selected_project:
        query["project_id"] = ObjectId(selected_project)

    shifts = list(mongo.db.shifts.find(query).sort("date", 1))
    
    # Convert shift _id and user_id to strings for template
    for s in shifts:
        s["_id"] = str(s["_id"])
        if s.get("user_id"):
            s["user_id"] = str(s["user_id"])
        if s.get("project_id"):
            s["project_id"] = str(s["project_id"])
    
    # Sort shifts by shift code in order: A, G, B, C
    shift_order = {"A": 1, "G": 2, "B": 3, "C": 4}
    shifts.sort(key=lambda x: (shift_order.get(x.get("shift_code", ""), 99), x.get("date", ""), x.get("start_time", "")))

    tasks = []
    if selected_project:
        tasks = list(
            mongo.db.project_tasks.find({"project_id": ObjectId(selected_project)}).sort(
                "created_at", -1
            )
        )

    all_members = list(mongo.db.users.find({"role": "member"}))
    users_map = {str(u["_id"]): u["name"] for u in all_members}
    projects_map = {str(p["_id"]): p["name"] for p in projects}

    shifts_map = {}
    if selected_date:
        cursor = mongo.db.shifts.find({"date": selected_date})
        for s in cursor:
            shifts_map[str(s["user_id"])] = s

    # Get today's date for date input min attribute
    from datetime import date
    today_date = date.today().isoformat()
    
    # Convert user _id to string for template (for bulk assignment checkboxes)
    for u in users:
        u["_id"] = str(u["_id"])
        if u.get("project_ids"):
            u["project_ids"] = [str(pid) for pid in u["project_ids"]]

    return render_template(
        "manager/manage_shifts.html",
        users=users,
        projects=projects,
        selected_date=selected_date,
        selected_project=selected_project,
        show_all_members=show_all_members,
        shifts_map=shifts_map,
        shifts=shifts,
        tasks=tasks,
        users_map=users_map,
        projects_map=projects_map,
        today_date=today_date,
    )


# =========================================================
# EXCEL UPLOAD FOR BULK SHIFT IMPORT
# =========================================================
@manager_bp.route("/upload-excel", methods=["GET", "POST"])
@manager_required  
def upload_excel():
    """
    Upload Excel file or copy-paste data to bulk import shifts.
    Expected columns: Date, Member Name, Project, Shift (A/B/C/G), Task
    Supports both Excel file upload and copy-paste from spreadsheet
    """
    if request.method == "POST":
        action = request.form.get("action", "upload")
        
        # Handle copy-paste import
        if action == "paste":
            paste_data = request.form.get("paste_data", "").strip()
            if not paste_data:
                flash("Please paste data from your spreadsheet.", "danger")
                return redirect("/manager/upload-excel")
            
            try:
                # Parse pasted data (tab or comma separated)
                lines = paste_data.strip().split('\n')
                if not lines:
                    flash("No data found in pasted content.", "danger")
                    return redirect("/manager/upload-excel")
                
                # Detect delimiter (tab or comma)
                first_line = lines[0]
                if '\t' in first_line:
                    delimiter = '\t'
                elif ',' in first_line:
                    delimiter = ','
                else:
                    # Try to split by multiple spaces
                    delimiter = None
                
                # Parse header row
                if delimiter:
                    headers = [h.strip().lower() for h in first_line.split(delimiter)]
                else:
                    headers = [h.strip().lower() for h in first_line.split()]
                
                # Find column indices
                date_col = None
                member_col = None
                project_col = None
                shift_col = None
                task_col = None
                
                for idx, header in enumerate(headers):
                    if "date" in header:
                        date_col = idx
                    elif "member" in header or ("name" in header and "project" not in header):
                        member_col = idx
                    elif "project" in header:
                        project_col = idx
                    elif "shift" in header:
                        shift_col = idx
                    elif "task" in header:
                        task_col = idx
                
                if date_col is None or member_col is None or shift_col is None:
                    flash("Pasted data must include Date, Member Name, and Shift columns.", "danger")
                    return redirect("/manager/upload-excel")
                
                # Get all users and projects for mapping
                users_list = list(mongo.db.users.find({"role": "member"}))
                users_map = {u.get("name", "").lower(): u["_id"] for u in users_list}
                users_map.update({u.get("email", "").lower(): u["_id"] for u in users_list})
                
                projects_list = list(mongo.db.projects.find())
                projects_map = {p.get("name", "").lower(): p["_id"] for p in projects_list}
                
                imported_count = 0
                error_count = 0
                errors = []
                
                # Process data rows (skip header)
                for row_idx, line in enumerate(lines[1:], 2):
                    try:
                        if not line.strip():
                            continue
                        
                        # Split row
                        if delimiter:
                            values = [v.strip() for v in line.split(delimiter)]
                        else:
                            values = [v.strip() for v in line.split()]
                        
                        # Get values
                        date_val = values[date_col] if date_col < len(values) else None
                        member_val = values[member_col] if member_col < len(values) else None
                        project_val = values[project_col] if project_col and project_col < len(values) else None
                        shift_val = values[shift_col] if shift_col < len(values) else None
                        task_val = values[task_col] if task_col and task_col < len(values) else None
                        
                        # Skip empty rows
                        if not date_val or not member_val or not shift_val:
                            continue
                        
                        # Parse date
                        date_str = None
                        date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"]
                        for fmt in date_formats:
                            try:
                                date_obj = datetime.strptime(date_val.strip(), fmt)
                                date_str = date_obj.strftime("%Y-%m-%d")
                                break
                            except:
                                continue
                        
                        if not date_str:
                            errors.append(f"Row {row_idx}: Invalid date format '{date_val}'")
                            error_count += 1
                            continue
                        
                        # Find user
                        member_key = str(member_val).strip().lower()
                        user_id = users_map.get(member_key)
                        if not user_id:
                            errors.append(f"Row {row_idx}: Member '{member_val}' not found")
                            error_count += 1
                            continue
                        
                        # Find project (optional)
                        project_id = None
                        if project_val and project_val.strip():
                            project_key = str(project_val).strip().lower()
                            project_id = projects_map.get(project_key)
                        
                        # Validate shift code
                        shift_code = str(shift_val).strip().upper()
                        if shift_code not in ["A", "B", "C", "G", "W", "L"]:
                            errors.append(f"Row {row_idx}: Invalid shift code '{shift_code}'. Must be A, B, C, G, W (Weekoff), or L (Leave)")
                            error_count += 1
                            continue
                        
                        # Get shift timings
                        start_time, end_time = SHIFT_TIMINGS.get(shift_code, ("09:00", "17:00"))
                        
                        # Check for existing shift
                        existing = mongo.db.shifts.find_one({
                            "user_id": user_id,
                            "date": date_str
                        })
                        
                        task_str = str(task_val).strip() if task_val else ""
                        
                        shift_doc = {
                            "user_id": user_id,
                            "date": date_str,
                            "shift_code": shift_code,
                            "start_time": start_time,
                            "end_time": end_time,
                            "task": task_str,
                            "project_id": project_id,
                            "created_at": datetime.utcnow(),
                            "updated_at": datetime.utcnow()
                        }
                        
                        if existing:
                            mongo.db.shifts.update_one(
                                {"_id": existing["_id"]},
                                {"$set": shift_doc}
                            )
                        else:
                            mongo.db.shifts.insert_one(shift_doc)
                        
                        # Send notification
                        mongo.db.notifications.insert_one({
                            "user_id": user_id,
                            "message": f"You have been assigned a {shift_code} shift on {date_str}.",
                            "created_at": datetime.utcnow(),
                            "read": False,
                        })
                        
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_idx}: {str(e)}")
                        error_count += 1
                        continue
                
                if imported_count > 0:
                    flash(f"Successfully imported {imported_count} shifts from pasted data! The calendar will refresh automatically.", "success")
                if error_count > 0:
                    flash(f"Encountered {error_count} errors during import. Check details below.", "warning")
                
                if errors:
                    session["excel_import_errors"] = errors[:20]  # Store first 20 errors
                
                # Redirect to dashboard to see the updated calendar
                if imported_count > 0:
                    return redirect(url_for("manager.dashboard") + "?imported=1")
                else:
                    return redirect("/manager/upload-excel")
                
            except Exception as e:
                flash(f"Error processing pasted data: {str(e)}", "danger")
                return redirect("/manager/upload-excel")
        
        # Handle Excel file upload
        if not OPENPYXL_AVAILABLE:
            flash("Excel upload feature requires 'openpyxl' package. Please install it using: pip install openpyxl", "danger")
            return redirect(url_for("manager.dashboard"))
        
        file = request.files.get("excel_file")
        
        if not file or file.filename == "":
            flash("Please select an Excel file.", "danger")
            return redirect("/manager/upload-excel")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash("Invalid file type. Please upload an Excel file (.xlsx or .xls).", "danger")
            return redirect("/manager/upload-excel")
        
        try:
            # Load workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Expected columns (case-insensitive matching)
            date_col = None
            member_col = None
            project_col = None
            shift_col = None
            task_col = None
            
            for idx, header in enumerate(headers, 1):
                header_lower = str(header).lower() if header else ""
                if "date" in header_lower:
                    date_col = idx
                elif "member" in header_lower or "name" in header_lower:
                    member_col = idx
                elif "project" in header_lower:
                    project_col = idx
                elif "shift" in header_lower:
                    shift_col = idx
                elif "task" in header_lower:
                    task_col = idx
            
            if not all([date_col, member_col, shift_col]):
                flash("Excel file must have Date, Member Name, and Shift columns.", "danger")
                return redirect("/manager/upload-excel")
            
            # Get all users and projects for mapping
            users_list = list(mongo.db.users.find({"role": "member"}))
            users_map = {u.get("name", "").lower(): u["_id"] for u in users_list}
            users_map.update({u.get("email", "").lower(): u["_id"] for u in users_list})
            
            projects_list = list(mongo.db.projects.find())
            projects_map = {p.get("name", "").lower(): p["_id"] for p in projects_list}
            
            imported_count = 0
            error_count = 0
            errors = []
            
            # Process rows (skip header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                try:
                    # Get values
                    date_val = row[date_col - 1].value if date_col <= len(row) else None
                    member_val = row[member_col - 1].value if member_col <= len(row) else None
                    project_val = row[project_col - 1].value if project_col and project_col <= len(row) else None
                    shift_val = row[shift_col - 1].value if shift_col <= len(row) else None
                    task_val = row[task_col - 1].value if task_col and task_col <= len(row) else None
                    
                    # Skip empty rows
                    if not date_val or not member_val or not shift_val:
                        continue
                    
                    # Parse date
                    date_str = None
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime("%Y-%m-%d")
                    elif isinstance(date_val, str):
                        date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]
                        for fmt in date_formats:
                            try:
                                date_obj = datetime.strptime(date_val.strip(), fmt)
                                date_str = date_obj.strftime("%Y-%m-%d")
                                break
                            except:
                                continue
                    
                    if not date_str:
                        errors.append(f"Row {row_idx}: Invalid date format '{date_val}'")
                        error_count += 1
                        continue
                    
                    # Find user
                    member_key = str(member_val).strip().lower()
                    user_id = users_map.get(member_key)
                    if not user_id:
                        errors.append(f"Row {row_idx}: Member '{member_val}' not found")
                        error_count += 1
                        continue
                    
                    # Find project (optional)
                    project_id = None
                    if project_val:
                        project_key = str(project_val).strip().lower()
                        project_id = projects_map.get(project_key)
                    
                    # Validate shift code
                    shift_code = str(shift_val).strip().upper()
                    if shift_code not in ["A", "B", "C", "G"]:
                        errors.append(f"Row {row_idx}: Invalid shift code '{shift_code}'. Must be A, B, C, or G")
                        error_count += 1
                        continue
                    
                    # Get shift timings
                    start_time, end_time = SHIFT_TIMINGS.get(shift_code, ("09:00", "17:00"))
                    
                    # Check for existing shift
                    existing = mongo.db.shifts.find_one({
                        "user_id": user_id,
                        "date": date_str
                    })
                    
                    task_str = str(task_val).strip() if task_val else ""
                    
                    shift_doc = {
                        "user_id": user_id,
                        "date": date_str,
                        "shift_code": shift_code,
                        "start_time": start_time,
                        "end_time": end_time,
                        "task": task_str,
                        "project_id": project_id,
                        "created_at": datetime.utcnow(),
                        "updated_at": datetime.utcnow()
                    }
                    
                    if existing:
                        mongo.db.shifts.update_one(
                            {"_id": existing["_id"]},
                            {"$set": shift_doc}
                        )
                    else:
                        mongo.db.shifts.insert_one(shift_doc)
                    
                    # Send notification
                    mongo.db.notifications.insert_one({
                        "user_id": user_id,
                        "message": f"You have been assigned a {shift_code} shift on {date_str}.",
                        "created_at": datetime.utcnow(),
                        "read": False,
                    })
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    error_count += 1
                    continue
            
            if imported_count > 0:
                flash(f"Successfully imported {imported_count} shifts from Excel! The calendar will refresh automatically.", "success")
            if error_count > 0:
                flash(f"Encountered {error_count} errors during import. Check details below.", "warning")
            
            if errors:
                session["excel_import_errors"] = errors[:20]  # Store first 20 errors
            
            # Redirect to dashboard to see the updated calendar
            if imported_count > 0:
                return redirect(url_for("manager.dashboard") + "?imported=1")
            else:
                return redirect("/manager/upload-excel")
            
        except Exception as e:
            flash(f"Error processing Excel file: {str(e)}", "danger")
            return redirect("/manager/upload-excel")
    
    # GET request - show upload form
    try:
        import_errors = session.pop("excel_import_errors", [])
    except:
        import_errors = []
    
    try:
        projects = list(mongo.db.projects.find().sort("name", 1))
        # Convert ObjectIds to strings for template
        for p in projects:
            p["_id"] = str(p["_id"])
    except Exception as e:
        current_app.logger.error(f"Error loading projects: {str(e)}")
        projects = []
    
    try:
        users = list(mongo.db.users.find({"role": "member"}).sort("name", 1))
        # Convert ObjectIds to strings for template
        for u in users:
            u["_id"] = str(u["_id"])
    except Exception as e:
        current_app.logger.error(f"Error loading users: {str(e)}")
        users = []
    
    return render_template("manager/upload_excel.html", errors=import_errors, projects=projects, users=users)


# =========================================================
# EDIT SHIFT (AUTO-TIMING ON UPDATE)
# =========================================================
@manager_bp.route("/edit-shift/<shift_id>", methods=["GET", "POST"], endpoint="edit_shift")
@manager_required
def edit_shift(shift_id):
    shift = mongo.db.shifts.find_one({"_id": ObjectId(shift_id)})
    if not shift:
        flash("Shift not found.", "danger")
        return redirect(url_for("manager.dashboard"))

    projects = list(mongo.db.projects.find())
    shift_project_id = shift.get("project_id")
    show_all = request.args.get("show_all", "false") == "true"

    if shift_project_id and not show_all:
        users = list(
            mongo.db.users.find(
                {"role": "member", "project_ids": shift_project_id}
            )
        )
        current_user_id = shift.get("user_id")
        if current_user_id:
            current_user = mongo.db.users.find_one({"_id": current_user_id})
            if current_user and current_user not in users:
                users.append(current_user)
    else:
        users = list(mongo.db.users.find({"role": "member"}))

    if request.method == "POST":
        date_str = request.form.get("date")
        user_id = request.form.get("user_id")
        shift_code = request.form.get("shift_code")
        task = request.form.get("task")
        project_id = request.form.get("project_id") or None

        start_time, end_time = SHIFT_TIMINGS.get(shift_code, ("09:00", "17:00"))

        has_conflict, _ = check_shift_conflict(
            user_id, date_str, exclude_shift_id=shift_id
        )
        if has_conflict:
            flash(
                f"Conflict: user already has a shift on {date_str}.",
                "danger",
            )
            return redirect(url_for("manager.edit_shift", shift_id=shift_id))

        update_doc = {
            "date": date_str,
            "user_id": ObjectId(user_id),
            "shift_code": shift_code,
            "start_time": start_time,
            "end_time": end_time,
            "task": task,
            "updated_at": datetime.utcnow(),
        }
        if project_id:
            update_doc["project_id"] = ObjectId(project_id)
        else:
            update_doc["project_id"] = None

        mongo.db.shifts.update_one(
            {"_id": ObjectId(shift_id)}, {"$set": update_doc}
        )

        mongo.db.notifications.insert_one(
            {
                "user_id": ObjectId(user_id),
                "message": f"Your shift on {date_str} has been updated by the manager.",
                "project_id": ObjectId(project_id) if project_id else None,
                "created_at": datetime.utcnow(),
                "read": False,
            }
        )

        flash("Shift updated and user notified.", "success")
        return redirect(url_for("manager.dashboard"))

    all_users = list(mongo.db.users.find({"role": "member"}))

    return render_template(
        "manager/edit_shift.html",
        shift=shift,
        users=users,
        all_users=all_users,
        projects=projects,
        shift_project_id=shift_project_id,
        show_all=show_all,
    )


# =========================================================
# DELETE SHIFT
# =========================================================
@manager_bp.route("/delete-shift/<shift_id>", methods=["POST"], endpoint="delete_shift")
@manager_required
def delete_shift(shift_id):
    shift = mongo.db.shifts.find_one({"_id": ObjectId(shift_id)})
    if not shift:
        flash("Shift not found.", "danger")
        return redirect(url_for("manager.manage_shifts"))

    user_id = shift.get("user_id")
    date_str = shift.get("date", "")

    mongo.db.shifts.delete_one({"_id": ObjectId(shift_id)})

    if user_id:
        mongo.db.notifications.insert_one(
            {
                "user_id": user_id,
                "message": f"Your shift on {date_str} has been removed.",
                "created_at": datetime.utcnow(),
                "read": False,
            }
        )

    flash("Shift deleted successfully.", "success")
    return redirect(url_for("manager.manage_shifts"))


# =========================================================
# REASSIGN SHIFT
# =========================================================
@manager_bp.route("/reassign-shift/<shift_id>", methods=["POST"], endpoint="reassign_shift")
@manager_required
def reassign_shift(shift_id):
    shift = mongo.db.shifts.find_one({"_id": ObjectId(shift_id)})
    if not shift:
        flash("Shift not found.", "danger")
        return redirect(url_for("manager.manage_shifts"))

    new_user_id = request.form.get("new_user_id")
    if not new_user_id:
        flash("Please select a user to reassign.", "danger")
        return redirect(url_for("manager.edit_shift", shift_id=shift_id))

    new_user_id = ObjectId(new_user_id)
    date_str = shift.get("date", "")
    old_user_id = shift.get("user_id")

    has_conflict, _ = check_shift_conflict(str(new_user_id), date_str, exclude_shift_id=shift_id)
    if has_conflict:
        user = mongo.db.users.find_one({"_id": new_user_id})
        user_name = user["name"] if user else "Unknown"
        flash(
            f"Conflict: {user_name} already has a shift on {date_str}.",
            "danger",
        )
        return redirect(url_for("manager.edit_shift", shift_id=shift_id))

    mongo.db.shifts.update_one(
        {"_id": ObjectId(shift_id)},
        {"$set": {"user_id": new_user_id, "updated_at": datetime.utcnow()}},
    )

    if old_user_id:
        mongo.db.notifications.insert_one(
            {
                "user_id": old_user_id,
                "message": f"Your shift on {date_str} has been reassigned.",
                "created_at": datetime.utcnow(),
                "read": False,
            }
        )

    mongo.db.notifications.insert_one(
        {
            "user_id": new_user_id,
            "message": f"You have been assigned a shift on {date_str}.",
            "created_at": datetime.utcnow(),
            "read": False,
        }
    )

    flash("Shift reassigned successfully.", "success")
    return redirect(url_for("manager.manage_shifts"))


# =========================================================
# CHANGE REQUESTS
# =========================================================
@manager_bp.route("/change-requests", methods=["GET", "POST"], endpoint="change_requests")
@manager_required
def change_requests():
    if request.method == "POST":
        req_id = request.form.get("req_id")
        action = request.form.get("action")

        if not req_id or not action:
            flash("Invalid request.", "danger")
            return redirect(url_for("manager.change_requests"))

        req = mongo.db.shift_change_requests.find_one({"_id": ObjectId(req_id)})
        if not req:
            flash("Request not found.", "danger")
            return redirect(url_for("manager.change_requests"))

        if action == "approve":
            new_code = req["requested_shift"]
            start_t, end_t = SHIFT_TIMINGS.get(new_code, ("09:00", "17:00"))

            mongo.db.shifts.update_one(
                {"date": req["date"], "user_id": req["user_id"]},
                {
                    "$set": {
                        "shift_code": new_code,
                        "start_time": start_t,
                        "end_time": end_t,
                        "updated_at": datetime.utcnow(),
                    }
                },
            )

            mongo.db.shift_change_requests.update_one(
                {"_id": ObjectId(req_id)},
                {"$set": {"status": "approved", "updated_at": datetime.utcnow()}},
            )

            mongo.db.notifications.insert_one(
                {
                    "user_id": req["user_id"],
                    "message": f"Your shift change request for {req['date']} has been approved.",
                    "created_at": datetime.utcnow(),
                    "read": False,
                }
            )

            flash("Request approved and shift updated.", "success")

        elif action == "reject":
            mongo.db.shift_change_requests.update_one(
                {"_id": ObjectId(req_id)},
                {"$set": {"status": "rejected", "updated_at": datetime.utcnow()}},
            )
            mongo.db.notifications.insert_one(
                {
                    "user_id": req["user_id"],
                    "message": f"Your shift change request for {req['date']} has been rejected.",
                    "created_at": datetime.utcnow(),
                    "read": False,
                }
            )
            flash("Request rejected.", "info")

        return redirect(url_for("manager.change_requests"))

    requests = list(
        mongo.db.shift_change_requests.find().sort("created_at", -1)
    )
    users_map = {str(u["_id"]): u["name"] for u in mongo.db.users.find()}

    return render_template(
        "manager/change_requests.html",
        requests=requests,
        users_map=users_map,
    )


# =========================================================
# SWAP REQUESTS
# =========================================================
@manager_bp.route("/swap-requests", methods=["GET", "POST"], endpoint="swap_requests")
@manager_required
def swap_requests():
    if request.method == "POST":
        req_id = request.form.get("req_id")
        action = request.form.get("action")

        if not req_id or not action:
            flash("Invalid request.", "danger")
            return redirect(url_for("manager.swap_requests"))

        req = mongo.db.shift_swap_requests.find_one({"_id": ObjectId(req_id)})
        if not req:
            flash("Request not found.", "danger")
            return redirect(url_for("manager.swap_requests"))

        if action == "approve":
            requester_shift = mongo.db.shifts.find_one(
                {"date": req["date"], "user_id": req["requester_id"]}
            )
            target_shift = mongo.db.shifts.find_one(
                {"date": req["date"], "user_id": req["target_user_id"]}
            )

            if requester_shift and target_shift:
                requester_shift_code = requester_shift.get("shift_code")
                target_shift_code = target_shift.get("shift_code")

                t_start, t_end = SHIFT_TIMINGS.get(target_shift_code, ("09:00", "17:00"))
                r_start, r_end = SHIFT_TIMINGS.get(requester_shift_code, ("09:00", "17:00"))

                mongo.db.shifts.update_one(
                    {"_id": requester_shift["_id"]},
                    {
                        "$set": {
                            "shift_code": target_shift_code,
                            "start_time": t_start,
                            "end_time": t_end,
                            "updated_at": datetime.utcnow(),
                        }
                    },
                )
                mongo.db.shifts.update_one(
                    {"_id": target_shift["_id"]},
                    {
                        "$set": {
                            "shift_code": requester_shift_code,
                            "start_time": r_start,
                            "end_time": r_end,
                            "updated_at": datetime.utcnow(),
                        }
                    },
                )

            mongo.db.shift_swap_requests.update_one(
                {"_id": ObjectId(req_id)},
                {"$set": {"status": "approved", "updated_at": datetime.utcnow()}},
            )

            mongo.db.notifications.insert_one(
                {
                    "user_id": req["requester_id"],
                    "message": f"Your shift swap request for {req['date']} has been approved.",
                    "created_at": datetime.utcnow(),
                    "read": False,
                }
            )
            mongo.db.notifications.insert_one(
                {
                    "user_id": req["target_user_id"],
                    "message": f"Your shift swap request for {req['date']} has been approved.",
                    "created_at": datetime.utcnow(),
                    "read": False,
                }
            )

            flash("Request approved and shifts swapped.", "success")

        elif action == "reject":
            mongo.db.shift_swap_requests.update_one(
                {"_id": ObjectId(req_id)},
                {"$set": {"status": "rejected", "updated_at": datetime.utcnow()}},
            )
            mongo.db.notifications.insert_one(
                {
                    "user_id": req["requester_id"],
                    "message": f"Your shift swap request for {req['date']} has been rejected.",
                    "created_at": datetime.utcnow(),
                    "read": False,
                }
            )
            flash("Request rejected.", "info")

        return redirect(url_for("manager.swap_requests"))

    requests = list(mongo.db.shift_swap_requests.find().sort("created_at", -1))
    users_map = {str(u["_id"]): u["name"] for u in mongo.db.users.find()}

    return render_template(
        "manager/swap_requests.html",
        requests=requests,
        users_map=users_map,
    )


# =========================================================
# LEAVE & WEEKOFF REQUESTS
# =========================================================
@manager_bp.route("/leave-requests", methods=["GET", "POST"], endpoint="leave_requests")
@manager_required
def leave_requests():
    if request.method == "POST":
        req_id = request.form.get("req_id")
        action = request.form.get("action")

        if not req_id or not action:
            flash("Invalid request.", "danger")
            return redirect(url_for("manager.leave_requests"))

        req = mongo.db.leave_requests.find_one({"_id": ObjectId(req_id)})
        if not req:
            flash("Request not found.", "danger")
            return redirect(url_for("manager.leave_requests"))

        if action == "approve":
            # Create shift entry for weekoff/leave
            shift_code = "W" if req["type"] == "weekoff" else "L"
            start_time, end_time = SHIFT_TIMINGS.get(shift_code, ("00:00", "23:59"))
            
            # Check if shift already exists
            existing_shift = mongo.db.shifts.find_one({
                "date": req["date"],
                "user_id": req["user_id"]
            })
            
            if existing_shift:
                # Update existing shift to weekoff/leave
                mongo.db.shifts.update_one(
                    {"_id": existing_shift["_id"]},
                    {
                        "$set": {
                            "shift_code": shift_code,
                            "start_time": start_time,
                            "end_time": end_time,
                            "updated_at": datetime.utcnow(),
                        }
                    }
                )
            else:
                # Create new shift entry
                mongo.db.shifts.insert_one({
                    "date": req["date"],
                    "user_id": req["user_id"],
                    "shift_code": shift_code,
                    "start_time": start_time,
                    "end_time": end_time,
                    "task": f"{req['type'].title()} - {req.get('reason', '')}",
                    "project_id": None,
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow(),
                })
            
            # Update request status
            mongo.db.leave_requests.update_one(
                {"_id": ObjectId(req_id)},
                {"$set": {"status": "approved", "updated_at": datetime.utcnow()}}
            )
            
            # Send notification
            mongo.db.notifications.insert_one({
                "user_id": req["user_id"],
                "message": f"Your {req['type']} request for {req['date']} has been approved.",
                "created_at": datetime.utcnow(),
                "read": False,
            })
            
            flash(f"{req['type'].title()} request approved and assigned.", "success")

        elif action == "reject":
            mongo.db.leave_requests.update_one(
                {"_id": ObjectId(req_id)},
                {"$set": {"status": "rejected", "updated_at": datetime.utcnow()}}
            )
            
            mongo.db.notifications.insert_one({
                "user_id": req["user_id"],
                "message": f"Your {req['type']} request for {req['date']} has been rejected.",
                "created_at": datetime.utcnow(),
                "read": False,
            })
            
            flash("Request rejected.", "info")

        return redirect(url_for("manager.leave_requests"))

    # Get all leave/weekoff requests
    requests = list(mongo.db.leave_requests.find().sort("created_at", -1))
    users_map = {str(u["_id"]): u["name"] for u in mongo.db.users.find()}

    return render_template(
        "manager/leave_requests.html",
        requests=requests,
        users_map=users_map,
    )


# =========================================================
# ASSIGN WEEKOFF/LEAVE DIRECTLY (MANAGER)
# =========================================================
@manager_bp.route("/assign-weekoff-leave", methods=["GET", "POST"], endpoint="assign_weekoff_leave")
@manager_required
def assign_weekoff_leave():
    if request.method == "POST":
        user_id = request.form.get("user_id")
        date_str = request.form.get("date")
        type_val = request.form.get("type")  # "weekoff" or "leave"
        reason = request.form.get("reason", "")
        
        if not user_id or not date_str or not type_val:
            flash("Please fill in all required fields.", "danger")
            return redirect(url_for("manager.assign_weekoff_leave"))
        
        if type_val not in ["weekoff", "leave"]:
            flash("Invalid type. Must be weekoff or leave.", "danger")
            return redirect(url_for("manager.assign_weekoff_leave"))
        
        shift_code = "W" if type_val == "weekoff" else "L"
        start_time, end_time = SHIFT_TIMINGS.get(shift_code, ("00:00", "23:59"))
        
        # Check for existing shift
        existing_shift = mongo.db.shifts.find_one({
            "date": date_str,
            "user_id": ObjectId(user_id)
        })
        
        if existing_shift:
            # Update existing shift
            mongo.db.shifts.update_one(
                {"_id": existing_shift["_id"]},
                {
                    "$set": {
                        "shift_code": shift_code,
                        "start_time": start_time,
                        "end_time": end_time,
                        "task": f"{type_val.title()} - {reason}",
                        "updated_at": datetime.utcnow(),
                    }
                }
            )
            flash(f"{type_val.title()} updated successfully.", "success")
        else:
            # Create new shift
            mongo.db.shifts.insert_one({
                "date": date_str,
                "user_id": ObjectId(user_id),
                "shift_code": shift_code,
                "start_time": start_time,
                "end_time": end_time,
                "task": f"{type_val.title()} - {reason}",
                "project_id": None,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow(),
            })
            flash(f"{type_val.title()} assigned successfully.", "success")
        
        # Send notification
        mongo.db.notifications.insert_one({
            "user_id": ObjectId(user_id),
            "message": f"You have been assigned {type_val} on {date_str}.",
            "created_at": datetime.utcnow(),
            "read": False,
        })
        
        return redirect(url_for("manager.assign_weekoff_leave"))
    
    # GET request - show form
    users = list(mongo.db.users.find({"role": "member"}).sort("name", 1))
    for u in users:
        u["_id"] = str(u["_id"])
    
    from datetime import date
    today_date = date.today().isoformat()
    
    return render_template(
        "manager/assign_weekoff_leave.html",
        users=users,
        today_date=today_date,
    )


# =========================================================
# MANAGER PROFILE (PROFILE PICTURE UPLOAD)
# =========================================================
# ---------------------------------------------------------
# MANAGER PROFILE PAGE (UPLOAD PROFILE PICTURE)
# ---------------------------------------------------------
@manager_bp.route("/profile", methods=["GET", "POST"], endpoint="profile")
@manager_required
def profile():
    user_id = ObjectId(session["user_id"])
    user = mongo.db.users.find_one({"_id": user_id})
    
    if not user:
        flash("User not found.", "danger")
        return redirect("/manager/dashboard")

    if request.method == "POST":
        file = request.files.get("profile_picture")

        if file and allowed_file(file.filename):
            filename = secure_filename(f"{user_id}.jpg")
            filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
            file.save(filepath)

            mongo.db.users.update_one(
                {"_id": user_id},
                {"$set": {"profile_picture": filename}}
            )

            flash("Profile picture updated!", "success")
            return redirect("/manager/profile")

        flash("Invalid image file type.", "danger")

    profile_pic = user.get("profile_picture", "default.png") if user else "default.png"
    return render_template("manager/profile.html", user=user, profile_pic=profile_pic)
